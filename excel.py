import openpyxl
from datetime import datetime


def export_excel(trello_client,
                 board_name: str,
                 list_ids: list,
                 list_names: list,
                 save_dir):
    """Excelファイルにエクスポート"""
    # excelの新規作成
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = 'list_name'
    sheet.cell(row=1, column=2).value = 'card_name'
    sheet.cell(row=1, column=3).value = 'desc'
    sheet.cell(row=1, column=4).value = 'due'

    for list_id, list_name in zip(list_ids, list_names):
        cards_in_list = trello_client.get_cards_in_list(list_id)
        for card in cards_in_list:
            # 書きこむ行を取得
            k = sheet.max_row + 1
            # カードの情報を書きこむ
            sheet.cell(row=k, column=1).value = list_name
            sheet.cell(row=k, column=2).value = card['name']
            sheet.cell(row=k, column=3).value = card['desc']
            sheet.cell(row=k, column=4).value = card['due']

    # 現在時刻から保存ファイル名を生成
    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_name = f'{now}_{board_name}.xlsx'
    wb.save(f'{save_dir}/{file_name}')
    return file_name


"""
あとでやる
def import_from_excel(trello_client, board_name):
    # 追加 ボード名が存在していない場合は新しく作る
    trello_board_names = trello_client.get_board_names()
    if board_name not in trello_board_names:
        trello_client.create_new_trello_board(board_name)

    board_id = trello_client.get_board_id(board_name)

    # ボードidから既に存在しているリスト名を取得しておく
    ids_and_names = trello_client.get_list_ids_and_names(board_id)
    trello_list_names = [list_name for list_id, list_name in ids_and_names]

    wb = openpyxl.load_workbook(f'{board_name}.xlsx')
    sheet = wb.active
    for i in range(2, sheet.max_row + 1):
        list_name = sheet.cell(row=i, column=1).value
        # リスト名がなかったら作成してリストを更新
        if list_name not in trello_list_names:
            trello_client.create_new_trello_list(board_id, list_name)
            trello_list_names.append(list_name)

        card_name = sheet.cell(row=i, column=2).value
        due_date = sheet.cell(row=i, column=3).value
        due_time = sheet.cell(row=i, column=4).value
        desc = sheet.cell(row=i, column=5).value
        list_id = trello_client.get_list_id(board_id, list_name)
        trello_client.add_task(list_id, card_name, due_date, due_time, desc)

    wb.close()
"""
